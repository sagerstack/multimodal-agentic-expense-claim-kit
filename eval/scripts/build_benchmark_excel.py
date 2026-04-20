"""Generate eval/BENCHMARKS.xlsx — one row per benchmark.

Run:
  poetry run python eval/scripts/build_benchmark_excel.py
"""

from pathlib import Path
import sys

_ROOT = Path(__file__).resolve().parents[2]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from eval.src.dataset import BENCHMARKS

# ---------------------------------------------------------------------------
# Static annotations per benchmark (not in TypedDict — editorial context)
# ---------------------------------------------------------------------------

ANNOTATIONS: dict[str, dict] = {
    "ER-001": {
        "image_contents": "GoRails / Example LLC receipt. Receipt #123, Mar 25 2024. Subscription $19.00, Tax $1.12, Total $20.12. Refund $5.00 on same day. Payment: ACH (garbled). Bill To: placeholder text.",
        "related_policy": "Documents must be receipts (not invoices or statements) to be accepted as expense proof.",
        "expected_reasoning": "Agent identifies document as a receipt, notes merchant/date/total are present. Should flag $5 refund as reducing net claimable amount. Must NOT invent expense category or employee identity.",
        "issues": "None — clean GoRails receipt. Note: same image as ER-018 (intentional — different questions).",
    },
    "ER-002": {
        "image_contents": "GoRails / Example LLC INVOICE. Invoice #123, Mar 25 2024. Status: PAID. Subscription $19.00, 0% tax, Amount Due $19.00. Header reads 'Invoice'.",
        "related_policy": "Only receipts are accepted as expense proof. Invoices must be routed to Accounts Payable.",
        "expected_reasoning": "Agent identifies document header as 'Invoice', not 'Receipt'. Rejects it as direct expense proof. Advises employee to submit to AP or obtain a payment receipt.",
        "issues": "Same image as 9.png / ER-009 — acceptable since both test invoice vs receipt distinction.",
    },
    "ER-003": {
        "image_contents": "GoRails / Example LLC STATEMENT. Statement #123, Mar 25 2024. Period: Feb 24 – Mar 25 2024. Subscription $19.00, 0% tax, Total $19.00. Header reads 'Statement'.",
        "related_policy": "Billing statements are not accepted as expense proof. A per-transaction receipt is required.",
        "expected_reasoning": "Agent identifies document as a 'Statement' covering a billing period. Rejects it and asks employee to obtain the underlying receipt.",
        "issues": "None — unique image, scenario matches.",
    },
    "ER-004": {
        "image_contents": "DIG restaurant, 150 East 52nd St New York NY. Date: May 28 2024. 1 Custom with Protein (Charred Chicken) $13.40. Subtotal $13.40, Tax $1.19, Tip $1.61, Total $16.20 USD. Payment: VISA CREDIT xxxxxxxx1417.",
        "related_policy": "A receipt must show at minimum: merchant name, transaction date, and total amount paid.",
        "expected_reasoning": "Agent extracts merchant (DIG), date (May 28 2024), and total ($16.20 USD). Confirms all three mandatory fields are legible. Marks receipt as complete.",
        "issues": "None — clean physical restaurant receipt. Also used for ER-014 (date window test).",
    },
    "ER-005": {
        "image_contents": "The Public Izakaya 2 (FK Synergo Pte Ltd), Jun 14 2024. 4 pax, Table 200. Items: Negi Toro Don $20.80, Yakitori Don $19.80, 2x Wagyu Sukiyaki $41.60. Subtotal $82.20, 10% Svc $8.22, 9% GST $8.14, Total $98.56 SGD. VISA.",
        "related_policy": "Extracted fields must match the receipt exactly. No inference of missing fields.",
        "expected_reasoning": "Agent extracts all fields accurately: merchant, date, each line item, subtotal, service charge, GST, total, currency, payment method.",
        "issues": "Also used for ER-008 (category) and ER-013 (duplicate detection). Different questions, defensible sharing.",
    },
    "ER-006": {
        "image_contents": "US Open SUV & Limo Service. Receipt #0138, Date: Jun 11 2024. Passenger: SAGAR. Pickup: 11E 48 St NYC. Destination: Newark. Car#13. Fare: $130 USD. Handwritten on pre-printed form.",
        "related_policy": "Extracted fields must match the receipt exactly. No inference of missing fields.",
        "expected_reasoning": "Agent extracts merchant, date, passenger name, pickup, destination, and fare. Notes payment type is not specified on the handwritten form — should not invent card type.",
        "issues": "Handwritten receipt — tests extraction robustness on non-digital format. Also used for ER-007.",
    },
    "ER-007": {
        "image_contents": "US Open SUV & Limo Service. Receipt #0138, Jun 11 2024. Fare: $130 USD. NYC to Newark airport transfer.",
        "related_policy": "Category taxonomy: Meals & Entertainment, Ground Transport / Parking, Air Travel, Accommodation, Fuel / Mileage Support, Office Supplies, Other.",
        "expected_reasoning": "Agent identifies this as a car service / limo transfer. Assigns 'Ground Transport / Parking' citing the car service company name and airport transfer route.",
        "issues": "None — clear category assignment. Same image as ER-006 (different questions).",
    },
    "ER-008": {
        "image_contents": "The Public Izakaya 2, Jun 14 2024. Japanese restaurant, 4 pax. Food items: Negi Toro Don, Yakitori Don, Wagyu Sukiyaki. Total $98.56 SGD.",
        "related_policy": "Category taxonomy: Meals & Entertainment, Ground Transport / Parking, Air Travel, Accommodation, Fuel / Mileage Support, Office Supplies, Other.",
        "expected_reasoning": "Agent identifies this as a restaurant meal with Japanese food items. Assigns 'Meals & Entertainment' citing restaurant name, food items, and dining context.",
        "issues": "Changed from original 'fuel receipt' benchmark — no fuel receipt available. Same image as ER-005 and ER-013.",
    },
    "ER-009": {
        "image_contents": "GoRails / Example LLC INVOICE. Invoice #123, Mar 25 2024. Status: PAID. Subscription $19.00. Header reads 'Invoice'.",
        "related_policy": "Vendor invoices must be routed to Accounts Payable, not processed as employee expense receipts.",
        "expected_reasoning": "Agent identifies this as a vendor invoice (not a personal expense receipt). Advises routing to AP. Does not process as employee reimbursement.",
        "issues": "Same image as 2.png / ER-002 — acceptable since both test invoice vs receipt routing.",
    },
    "ER-010": {
        "image_contents": "Cari Truong restaurant, Hanoi Vietnam. Date: Aug 16 2018. 6 line items in VND: Bia tuoi (beer) x5 150,000, Pepper beef rice 65,000, Fried spring rolls 65,000, French fries 60,000, Coke 20,000, 'weed' 150,000. Total: 510,000 VND.",
        "related_policy": "Foreign currency amounts must be captured as-printed. Do not convert at extraction stage. Tax lines must be extracted individually.",
        "expected_reasoning": "Agent extracts currency=VND, total=510,000, and all 6 line items with quantities and unit prices. Does NOT convert to SGD at this stage. May note 'weed' item as suspicious.",
        "issues": "Contains 'weed' as a line item — ambiguous. Also used for ER-020 (cross-receipt consistency).",
    },
    "ER-011": {
        "image_contents": "SERVOS German Burger Grill, VivoCity. Mar 27 2025. 20+ line items: Beer Bucket x2 (78.00), Chili Cheese x7 (63.00), Fries x3 (21.00), Sausage Platter x3 (86.70), and many more. Subtotal 606.40, Svc 60.64, GST 60.05, Grand Total 727.09 SGD.",
        "related_policy": "Receipts with distinct line items should be flagged as itemizable for per-category splitting.",
        "expected_reasoning": "Agent detects 20+ distinct line items on the receipt. Confirms receipt structure supports per-item expense splitting. Lists the individual items or item categories.",
        "issues": "Also used for ER-015, ER-016, ER-017, ER-019. Different questions on same receipt.",
    },
    "ER-012": {
        "image_contents": "GoRails / Example LLC receipt. Mar 25 2024, Subscription $20.12 USD. (Same image as 1.png / ER-001.)",
        "related_policy": "Receipt must match expense entry on date (±1 day), amount (exact), and vendor name.",
        "expected_reasoning": "Expense entry: Contoso, Jun 10 2024, $2,594.28 USD. Receipt: Example LLC, Mar 25 2024, $20.12 USD. Date, amount, and vendor all differ — clear mismatch on all three fields.",
        "issues": "Intentional mismatch scenario — GoRails receipt is obviously wrong for the Contoso expense entry. Expected decision is 'No Match'.",
    },
    "ER-013": {
        "image_contents": "The Public Izakaya 2, Jun 14 2024, $98.56 SGD. SECOND SUBMISSION of the same receipt as ER-005.",
        "related_policy": "Receipts matching a prior submission on merchant, date, and total must be flagged as duplicate risk.",
        "expected_reasoning": "Agent compares current receipt (Izakaya, Jun 14 2024, $98.56 SGD) against prior receipt history. Finds exact match on merchant, date, and total. Flags as duplicate risk.",
        "issues": "Playwright runner must submit this receipt twice in sequence (two-session pattern).",
    },
    "ER-014": {
        "image_contents": "DIG restaurant, NYC. Date: May 28 2024, Total $16.20 USD.",
        "related_policy": "Receipts must be submitted within 30 days of purchase date. Submission date: Sep 15 2024.",
        "expected_reasoning": "Agent reads purchase date (May 28 2024). Computes elapsed time to submission date (Sep 15 2024) = 110 days. Exceeds 30-day window. Flags as late submission requiring manager approval.",
        "issues": "None — date arithmetic is deterministic: May 28 to Sep 15 = 110 days. Clear late submission.",
    },
    "ER-015": {
        "image_contents": "SERVOS German Burger Grill. Grand Total: 727.09 SGD. Claimed amount in report: 780.60 SGD.",
        "related_policy": "Receipt total must match the claimed amount exactly. Any discrepancy flags the line for review.",
        "expected_reasoning": "Agent extracts Grand Total 727.09 SGD from receipt. Compares to claimed 780.60 SGD. Reports discrepancy of 53.51 SGD. Flags as Amount Mismatch.",
        "issues": "None — deliberate mismatch: claimed 780.60 vs actual 727.09.",
    },
    "ER-016": {
        "image_contents": "SERVOS German Burger Grill. Grand Total: 727.09 SGD.",
        "related_policy": "Standard: <500 SGD → direct manager. Senior: 500–2000 SGD → department head. Finance: >2000 SGD → finance controller.",
        "expected_reasoning": "Agent extracts total 727.09 SGD. Applies threshold: 727.09 is between 500 and 2000 SGD. Routes to Senior Approval (department head).",
        "issues": "None — threshold arithmetic is clear: 727.09 falls in 500–2000 range.",
    },
    "ER-017": {
        "image_contents": "SERVOS German Burger Grill. Items include: 2x Beer Bucket of 5 (78.00), 2x HH Beer (20.00), Cranberry Weissbier, Lager Beer. Total 727.09 SGD.",
        "related_policy": "Alcohol is not reimbursable. Meals cap: SGD 40 local / SGD 80 international. Entertainment >SGD 200 requires pre-approval.",
        "expected_reasoning": "Agent identifies alcohol line items (Beer Bucket, HH Beer, Lager Beer, Weissbier) and cites policy that alcohol is not reimbursable. Marks as Out of Policy. May also flag total 727.09 > SGD 200 entertainment threshold.",
        "issues": "None — alcohol items are explicit on the receipt.",
    },
    "ER-018": {
        "image_contents": "GoRails / Example LLC receipt. Receipt #123, Mar 25 2024. Subscription $19.00, Tax $1.12, Total $20.12. Refund $5.00 on Mar 25 2024. Payment: 'ACH super long super long super long super long super long' (garbled). Bill To: placeholder text (Customer, Address, City State Zipcode).",
        "related_policy": "Agents must not invent expense categories, employee IDs, or fields not visible on the receipt.",
        "expected_reasoning": "Agent reports Total $20.12, Refund $5.00 and flags net amount as ambiguous. Asks for clarification. Does NOT invent: expense category, employee name, project code, or claimant identity. Flags garbled payment method as unreadable.",
        "issues": "Same image as ER-001 — intentional. ER-018 tests hallucination avoidance on the same receipt that ER-001 classifies.",
    },
    "ER-019": {
        "image_contents": "Carrefour receipt. Photographed at 90° angle, heavily blurred. Merchant logo (Carrefour) partially visible. All text fields — date, items, amounts, total — are unreadable.",
        "related_policy": "Receipts with OCR confidence below 85% or critical fields unreadable must be escalated for human review.",
        "expected_reasoning": "Agent attempts image quality check (blur detection, resolution). Determines critical fields are unreadable. Escalates to human reviewer — does not attempt to extract or invent field values.",
        "issues": "None — genuinely unreadable blurry scan. Clear escalation case.",
    },
    "ER-020": {
        "image_contents": "Cari Truong, Hanoi Vietnam. Aug 16 2018. 510,000 VND. Vietnamese food and 'weed' line item.",
        "related_policy": "All receipts in a report must be consistent with the stated trip purpose, dates, and total claimed amount.",
        "expected_reasoning": "Report: Contoso client visit, Redmond WA, Jun 10–12 2024, USD. Receipt: Hanoi Vietnam, Aug 2018, VND. Agent flags at least one inconsistency: country mismatch, date mismatch, or currency mismatch.",
        "issues": "Three clear inconsistencies make this a strong test. Note: also raises non-reimbursable 'weed' item as secondary concern.",
    },
}

# ---------------------------------------------------------------------------
# Column layout
# ---------------------------------------------------------------------------

COLUMNS = [
    ("ID", 8),
    ("Benchmark", 30),
    ("Category", 14),
    ("File", 22),
    ("Image Contents", 50),
    ("Scenario", 40),
    ("Question", 38),
    ("Related Policy", 45),
    ("Expected Decision", 28),
    ("Expected Reasoning", 55),
    ("Pass Criteria", 45),
    ("Scoring Type", 18),
    ("Issues / Notes", 50),
]

CAT_FILLS = {
    "classification": PatternFill("solid", fgColor="D6E4FF"),
    "extraction":     PatternFill("solid", fgColor="D6F5D6"),
    "reasoning":      PatternFill("solid", fgColor="FFF3CC"),
    "workflow":       PatternFill("solid", fgColor="F9DDFF"),
    "safety":         PatternFill("solid", fgColor="FFD6D6"),
}

CRITICAL_FILL = PatternFill("solid", fgColor="FFB3B3")
WARN_FILL     = PatternFill("solid", fgColor="FFE0B2")
OK_FILL       = PatternFill("solid", fgColor="C8E6C9")
HEADER_FILL   = PatternFill("solid", fgColor="2E4057")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
BOLD          = Font(bold=True)


def _issue_fill(text: str) -> PatternFill | None:
    t = text.lower()
    if "needs improvement" in t or "critical" in t:
        return CRITICAL_FILL
    if "also used" in t or "same image" in t:
        return WARN_FILL
    return OK_FILL


def build_excel(output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Benchmarks"

    # Header
    for ci, (name, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for ri, b in enumerate(BENCHMARKS, 2):
        ann = ANNOTATIONS.get(b["benchmarkId"], {})
        cat_fill = CAT_FILLS.get(b["category"])
        issues_text = ann.get("issues", "")

        row_values = [
            b["benchmarkId"],
            b["benchmark"],
            b["category"],
            b["file"],
            ann.get("image_contents", ""),
            b["scenario"],
            b["question"],
            ann.get("related_policy", ""),
            b["expectedDecision"],
            ann.get("expected_reasoning", ""),
            b["passCriteria"],
            b["scoringType"],
            issues_text,
        ]

        for ci, value in enumerate(row_values, 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if ci <= 4 and cat_fill:
                cell.fill = cat_fill
            if ci == 1:
                cell.font = BOLD
            if ci == len(row_values):
                cell.fill = _issue_fill(issues_text)

        ws.row_dimensions[ri].height = 90

    # Image map sheet
    ws2 = wb.create_sheet("Image Map")
    map_cols = [("File", 28), ("Used By", 35), ("What it is", 55), ("Status", 12)]
    for ci, (name, width) in enumerate(map_cols, 1):
        c = ws2.cell(row=1, column=ci, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.column_dimensions[get_column_letter(ci)].width = width

    image_map = [
        ("1.png", "ER-001, ER-018", "GoRails / Example LLC — Receipt. Refund + garbled payment field.", "OK"),
        ("2.png", "ER-002", "GoRails / Example LLC — Invoice (PAID). Header: 'Invoice'.", "OK"),
        ("3.png", "ER-003", "GoRails / Example LLC — Statement. Period billing.", "OK"),
        ("9.png", "ER-009", "GoRails / Example LLC — Invoice (same as 2.png). Different question.", "OK"),
        ("12.png", "ER-012", "GoRails / Example LLC — Receipt (same as 1.png). Intentional mismatch test.", "OK"),
        ("dig-restaurant.jpeg", "ER-004, ER-014", "DIG restaurant NYC. $16.20 USD, May 28 2024. Clean, complete.", "OK"),
        ("izakaya-public.jpeg", "ER-005, ER-008, ER-013", "The Public Izakaya 2. $98.56 SGD, Jun 14 2024. Itemized.", "OK"),
        ("limo-usopensuv.jpeg", "ER-006, ER-007", "US Open SUV & Limo. $130 USD, Jun 11 2024. Handwritten.", "OK"),
        ("german-burger-servos.jpeg", "ER-011, ER-015, ER-016, ER-017", "SERVOS German Burger. 727.09 SGD, Mar 27 2025. Alcohol. 20+ items.", "Review"),
        ("vietnamese-cari-truong.jpg", "ER-010, ER-020", "Cari Truong, Hanoi. 510,000 VND, Aug 2018. 'Weed' item.", "OK"),
        ("receipt-blurry.jpg", "ER-019", "Carrefour receipt. Rotated 90°, heavily blurred. All text fields unreadable.", "OK"),
    ]

    status_fills = {
        "OK": PatternFill("solid", fgColor="C8E6C9"),
        "Review": PatternFill("solid", fgColor="FFE0B2"),
        "Critical": PatternFill("solid", fgColor="FFB3B3"),
    }

    for ri, (file, benchmarks, desc, status) in enumerate(image_map, 2):
        ws2.cell(row=ri, column=1, value=file).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.cell(row=ri, column=2, value=benchmarks).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.cell(row=ri, column=3, value=desc).alignment = Alignment(wrap_text=True, vertical="top")
        sc = ws2.cell(row=ri, column=4, value=status)
        sc.alignment = Alignment(horizontal="center", vertical="top")
        sc.fill = status_fills.get(status, OK_FILL)
        ws2.row_dimensions[ri].height = 45

    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"  {len(BENCHMARKS)} benchmarks across {len(set(b['category'] for b in BENCHMARKS))} categories")


if __name__ == "__main__":
    output = _ROOT / "eval" / "BENCHMARKS.xlsx"
    build_excel(output)
